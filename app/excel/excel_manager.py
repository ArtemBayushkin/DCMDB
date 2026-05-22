import os
import sys

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import QApplication, QMainWindow, QFileDialog, QPushButton, QWidget, QVBoxLayout

CQ_COLUMN_MAP: dict[str, str] = {
    "ID": "ID",
    "Date of meeting": "Date of meeting",
    "Code of the working document": "Code of the WD or MD",
    "Description of problem": "Description of problem",
    "Symbols of decisions under the Protocol": "Symbols of decisions under the Protocol",
    "Texts of  decisions, date": "Texts of  decisions, date",
    "Texts of decisions, date": "Texts of  decisions, date",
    "Desighners surname": "Desighner's surname",
    "Desighner's surname": "Desighner's surname",
    "Designers surname": "Desighner's surname",
    "Designer's surname": "Desighner's surname",
    "提出人": "От кого вопрос",
}

CQ_FIXED_VALUES: dict[str, object] = {
    "Urgent/Срочный": True,
    "Вопрос в рабочем порядке": True
}

# Обычные вопросы (DCM / regular)
DCM_COLUMN_MAP: dict[str, str] = {
    "ID": "ID",
    "Date of meeting": "Date of meeting",
    "Code of the DDD/MD to which the issue pertains": "Code of the WD or MD",
    "Description of problem": "Description of problem",
    "Symbols of decisions under the Protocol": "Symbols of decisions under the Protocol",
    "Texts of  decisions, date": "Texts of  decisions, date",
    "Texts of decisions, date": "Texts of  decisions, date",
    "Desighners surname": "Desighner's surname",
    "Desighner's surname": "Desighner's surname",
    "Designers surname": "Desighner's surname",
    "Designer's surname": "Desighner's surname",
    "Whether it is urgent (Yes or No)": "Urgent/Срочный",
    "问题提出者": "От кого вопрос",
}
DCM_FIXED_VALUES: dict[str, object] = {}


class ExcelParser(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Parser")
        self.setGeometry(100, 100, 300, 200)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        self.button = QPushButton('Open File')
        self.button.clicked.connect(ExcelParser.open_file)
        layout.addWidget(self.button)

    @staticmethod
    def read_two_row_header_excel(file_path: list) -> pd.DataFrame:
        """
        Читает Excel-файлы с двухстрочным заголовком (формат CQ и DCM).

        Строка 0 — верхний заголовок: ID, Date of meeting, Code..., Description..., Protocol, ...
        Строка 1 — нижний заголовок (подколонки): Symbols..., Texts..., Designer's surname, ...
        Строки 2+ — данные.

        Для каждой колонки берёт строку 0 если непустая, иначе строку 1.

        :param file_path: список путей к Excel-файлам
        :return: объединённый DataFrame с корректными именами колонок
        """
        print("ExcelParser.read_two_row_header_excel -> start")
        all_dfs = []
        for file in file_path:
            for sheet_name in pd.ExcelFile(file).sheet_names:
                raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
                if len(raw) < 3:
                    continue

                row0 = raw.iloc[0].fillna("").astype(str).str.strip()
                row1 = raw.iloc[1].fillna("").astype(str).str.strip()

                # Берём имя из строки 0; если пустое — из строки 1
                columns = [r0 if r0 else r1 for r0, r1 in zip(row0, row1)]

                df = raw.iloc[2:].copy()
                df.columns = columns
                df = df.reset_index(drop=True)
                df = df.dropna(how="all")

                if not df.empty:
                    print(f"read_two_row_header_excel: {file} / {sheet_name} "
                          f"-> {len(df)} rows, cols: {list(df.columns)}")
                    all_dfs.append(df)

        if all_dfs:
            return pd.concat(all_dfs, ignore_index=True)
        return pd.DataFrame()

    @staticmethod
    def read_excel(file_path: list, **kwargs) -> pd.DataFrame:
        """
        Чтение данных из Excel файлов (однострочный заголовок, skiprows=1).
        Оставлен для обратной совместимости.
        Для файлов CQ/DCM используйте read_two_row_header_excel.

        :param file_path: список путей к файлам
        :return: DataFrame с объединёнными данными всех листов
        """
        print("ExcelParser.read_excel -> start")
        all_dfs = []
        for file in file_path:
            for sheet_name in pd.ExcelFile(file).sheet_names:
                df = pd.read_excel(file, sheet_name=sheet_name, skiprows=1, **kwargs)
                if not df.empty:
                    all_dfs.append(df)
        if all_dfs:
            return pd.concat(all_dfs, ignore_index=True)
        return pd.DataFrame()

    @staticmethod
    def open_folder(start_path: str | None = None) -> str | None:
        """
        Диалог выбора папки с файлами приложений (Appendix / Приложение).

        :param start_path: начальная директория (по умолчанию None)
        :return: путь к папке в формате UNC/Windows или None при отмене
        """
        folder = QFileDialog.getExistingDirectory(
            None,
            "Выберите папку с файлами приложений",
            start_path or "",
        )
        if folder:
            return os.path.abspath(folder)
        else:
            return None

    @staticmethod
    def build_hyperlinks(df: pd.DataFrame, folder_path: str, mode: str) -> pd.DataFrame:
        """
        Заполняет колонки ``Appendix`` и ``Приложение`` гиперссылками
        в формате Access (#путь#).

        Правила формирования пути:

        **mode='cq'** (срочные вопросы):
            - Appendix    = ``#{folder}\{ID}.pdf#``
            - Приложение  = ``#{folder}\Appendices#``

        **mode='dcm'** (обычные вопросы):
            - Appendix    = ``#{folder}\ID{ID}.pdf#``
            - Приложение  = ``#{folder}\#``

        :param df:          DataFrame с колонкой ``ID`` (уже нормализованный)
        :param folder_path: путь к папке (из open_folder)
        :param mode:        ``'cq'`` или ``'dcm'``
        :return: копия DataFrame с заполненными гиперссылками
        """
        # Нормализуем разделители → обратные слеши, убираем trailing slash
        folder = folder_path.replace("/", "\\").rstrip("\\")

        df = df.copy()
        appendix_vals = []
        prilozhenie_vals = []

        for _, row in df.iterrows():
            id_val = str(row.get("ID", "")).strip()

            if mode == "cq":
                appendix = f"#{folder}\\{id_val}.pdf#"
                prilozhenie = f"#{folder}\\Appendices#"
            else:  # dcm
                appendix = f"#{folder}\\ID{id_val}.pdf#"
                prilozhenie = f"#{folder}\\Appendices#"

            appendix_vals.append(appendix)
            prilozhenie_vals.append(prilozhenie)

        df["Appendix"] = appendix_vals
        df["Приложение"] = prilozhenie_vals
        return df

    @staticmethod
    def write_excel(df, selected_columns: list[str] | None = None, column_widths: dict[str, int] | None = None)->bool:
        """
        Записывает DataFrame в Excel-файл с настройками ширины столбцов и выравнивания.
        """
        file_path, _ = QFileDialog.getSaveFileName(
            caption='Выберите место сохранения файла',
            filter='Excel Files (*.xlsx)'
        )
        if not file_path:
            return False

        print("ExcelParser.write_excel -> file_path", file_path)

        # Получаем DataFrame
        if hasattr(df, '_dataframe'):
            df = df._dataframe

        # Столбцы по умолчанию
        if selected_columns is None:
            selected_columns = [
                "ID", "Date of meeting", "Code of the WD or MD",
                "Description of problem", "Symbols of decisions under the Protocol",
                "Texts of  decisions, date", "Desighner's surname", "От кого вопрос"
            ]

        # Ширина в Excel единицах
        if column_widths is None:
            column_widths = {
                "ID": 15.2, "Date of meeting": 13.9, "Code of the WD or MD": 23.3,
                "Description of problem": 51.4, "Symbols of decisions under the Protocol": 18.0,
                "Texts of  decisions, date": 43.6, "Desighner's surname": 13.6, "От кого вопрос": 13.9
            }

        available_columns = [col for col in selected_columns if col in df.columns]
        new_df = df[available_columns].copy()

        if "Date of meeting" in new_df.columns:
            new_df["Date of meeting"] = pd.to_datetime(new_df["Date of meeting"], errors='coerce')
            new_df["Date of meeting"] = new_df["Date of meeting"].dt.strftime('%d.%m.%Y')
            new_df["Date of meeting"] = new_df["Date of meeting"].fillna('')

        if "ID" in new_df.columns:
            new_df = new_df.sort_values(by=['ID'])

        try:
            new_df.to_excel(file_path, index=False)

            wb = load_workbook(file_path)
            ws = wb.active

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            for col_idx, col_name in enumerate(available_columns, start=1):
                width = column_widths.get(col_name, 10)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            print(f"Excel файл успешно сохранён: {file_path}")
            return True

        except ImportError:
            new_df.to_excel(file_path, index=False)
            print("Excel файл сохранён без форматирования (установите openpyxl)")
            return True
        except Exception as e:
            print(f"Ошибка при сохранении Excel: {e}")
            return False

    @staticmethod
    def open_file(path=None):
        """
        Диалог выбора Excel-файлов.

        :param path: начальная директория
        :return: (список файлов, папка первого файла) или None при отмене
        """
        try:
            file_name, _ = QFileDialog.getOpenFileNames(
                parent=None,
                caption='Выберите Excel файлы с формой вопроса',
                directory=path,
                filter='Excel Files (*.xlsx)'
            )
            if file_name:
                folder_path = os.path.abspath(file_name[0])
                return file_name, folder_path
            return None
        except Exception as e:
            print(f"ExcelParser.open_file -> ошибка: {e}")
            return None

    # ──────────────────────────────────────────────────────────────
    # Нормализация DataFrame под UPLOAD_COLUMNS
    # ──────────────────────────────────────────────────────────────

    @staticmethod
    def normalize_dataframe(
            df: pd.DataFrame,
            column_map: dict[str, str],
            fixed_values: dict[str, object] | None = None,
            target_columns: list[str] | None = None,
    ) -> pd.DataFrame:
        """
        Приводит произвольный DataFrame к целевой схеме колонок.

        Алгоритм:
        1. Переименовывает колонки по ``column_map`` (source → target).
        2. Добавляет колонки с фиксированными значениями из ``fixed_values``.
        3. Если передан ``target_columns`` — оставляет только эти колонки
           в указанном порядке; отсутствующие заполняет None.

        :param df:             исходный DataFrame (после read_excel)
        :param column_map:     словарь {имя_в_excel: целевое_имя}
        :param fixed_values:   словарь {целевая_колонка: значение} — значения,
                               которых нет в Excel (например Urgent/Срочный=True)
        :param target_columns: список целевых колонок в нужном порядке
        :return: нормализованный DataFrame
        """
        # Переименовываем только те колонки, которые реально есть в df
        rename_map = {src: dst for src, dst in column_map.items() if src in df.columns}
        df = df.rename(columns=rename_map)

        # Фиксированные значения
        if fixed_values:
            for col, val in fixed_values.items():
                df[col] = val

        # Приводим к целевому набору колонок
        if target_columns:
            for col in target_columns:
                if col not in df.columns:
                    df[col] = None
            df = df[target_columns]

        return df

    # ──────────────────────────────────────────────────────────────
    # Парсинг шифров и подбор специалистов
    # ──────────────────────────────────────────────────────────────

    @staticmethod
    def parse_text(row):
        """
        Извлекает специальности, здания и системы из шифра документа.

        :param row: строка с шифром
        :return: (специальности, здания, системы) — множества
        """
        spis = re.findall(r'(?:LYG|XDP)-.*?(?=(?:LYG|XDP)-|$)', row)
        print(f"parse_text -> spis: {spis}")
        specialities = set()
        buildings = set()
        systems = set()

        for ite, i in enumerate(spis):
            print(f'\nИтерация № {ite}')
            f = i.split('-')[1:]
            print(f'  Разбивка: {f}')
            specialities.add(f[1])
            buildings.add(f[3][1:4])
            systems.add(f[3][6:9])
            print(f'  spec={f[1]}, building={f[3][1:4]}, system={f[3][6:9]}')

        return specialities, buildings, systems

    @staticmethod
    def find_best_specialist(specialities, buildings, systems, emp_df):
        """
        Определяет подходящего сотрудника по 2 из 3 критериев
        (специальность, здание, система).

        :return: строка вида «A. Ivanov / B. Petrov» или пустая строка
        """
        best_matches = []
        print(f"find_best_specialist -> emp_df.shape = {emp_df.shape}")
        print(f"find_best_specialist -> columns: {list(emp_df.columns)}")

        for idx, row in emp_df.iterrows():
            score = 0
            matches = {'speciality': False, 'building': False, 'system': False}

            if pd.notna(row['Код_специальности']):
                emp_specs = [s.strip() for s in str(row['Код_специальности']).split(',')]
                for spec in specialities:
                    if spec in emp_specs:
                        score += 3
                        matches['speciality'] = True
                        break

            if row['Перечень_зданий'] == 'Все здания':
                score += 2
                matches['building'] = True
            elif pd.notna(row['Перечень_зданий']):
                emp_buildings = [b.strip() for b in str(row['Перечень_зданий']).split(',')]
                for building in buildings:
                    if building in emp_buildings:
                        score += 2
                        matches['building'] = True
                        break

            if row['Перечень_систем_по_всем_зданиям'] == 'Все здания':
                score += 1
                matches['system'] = True
            elif pd.notna(row['Перечень_систем_по_всем_зданиям']):
                emp_systems = re.findall(r'[A-Z]{3}', str(row['Перечень_систем_по_всем_зданиям']))
                for system in systems:
                    if system in emp_systems:
                        score += 1
                        matches['system'] = True
                        break

            if matches['system'] + matches['building'] + matches['speciality'] >= 2:
                best_matches.append(row['ФИО_англ'])
        print(f"find_best_specialist -> {best_matches}")
        return ' / '.join(best_matches)

    def create_multiple_specialists_df(self, df: pd.DataFrame, emp: pd.DataFrame, num: int = 1) -> pd.DataFrame:
        """
        Добавляет в DataFrame колонку с фамилиями специалистов,
        определённых по шифру документа.

        :param df:  исходный DataFrame (из read_excel)
        :param emp: DataFrame сотрудников
        :param num: индекс колонки с шифром документа
                    (1 — для обычных DCM, 2 — для CQ)
        :return: DataFrame с заполненной колонкой «Designer's surname»
        """
        results = []
        for row in df.iloc[:, num]:
            print(f"\nОбработка: {row}")
            specialities, buildings, systems = self.parse_text(str(row))
            print(f"  spec={specialities}, bld={buildings}, sys={systems}")
            matches = self.find_best_specialist(specialities, buildings, systems, emp)
            print(f"  → {matches}")
            results.append(matches)

        # Записываем в ту колонку, которая уже есть, или создаём новую
        target_col = "Designer's surname" if "Designer's surname" in df.columns else "Desighner's surname"
        df[target_col] = results
        print(df.to_string())
        return df


# ──────────────────────────────────────────────────────────────────────────────
# Точка входа для ручного тестирования
# ──────────────────────────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    window = ExcelParser()
    result = window.open_file()
    if result is None:
        print("Файлы не выбраны")
        return
    file_list, file_path = result
    print(f'file_list: {file_list}\nfile_path: {file_path}')
    df = window.read_excel(file_list)
    print(f"Прочитано строк: {len(df)}")
    window.write_excel(df)


if __name__ == "__main__":
    main()
